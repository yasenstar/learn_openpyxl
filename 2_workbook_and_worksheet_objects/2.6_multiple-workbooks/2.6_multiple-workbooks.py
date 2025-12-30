from openpyxl import Workbook, load_workbook

workbook1 = load_workbook("my_workbook.xlsx")
workbook2 = Workbook()

print(workbook1.sheetnames)

# some processing
workbook1.remove(workbook1.worksheets[3])
new_sheet = workbook2.create_sheet("new sheet")

print(workbook1.sheetnames)
print(workbook1.properties)

workbook1.save("my_workbook_modified.xlsx")
workbook2.save("file2.xlsx")