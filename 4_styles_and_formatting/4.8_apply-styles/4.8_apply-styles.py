from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

workbook = Workbook()
sheet = workbook.active

# Define styles once
bold_red_font = Font(bold = True, color = "FF0000")
yellow_fill = PatternFill(
    start_color = "FFFF00",
    end_color = "FFFF00",
    fill_type = "solid"
)
thin_border = Border(
    left = Side(style = "thin"),
    right = Side(style = "thin"),
    top = Side(style = "thin"),
    bottom = Side(style = "thick")
)
center_alignment = Alignment(horizontal="center")

# Define styles to a range
for row in sheet.iter_rows(min_row=1, max_row=3, min_col=1, max_col=3):
    for cell in row:
        cell.value = "hello"
        cell.font = bold_red_font
        cell.fill = yellow_fill
        cell.border = thin_border
        cell.alignment = center_alignment

workbook.save("range_styles.xlsx")