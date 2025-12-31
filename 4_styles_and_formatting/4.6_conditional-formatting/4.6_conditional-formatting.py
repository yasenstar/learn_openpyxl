from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Color
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

wb = Workbook()
ws = wb.active

# Create fill
redFill = PatternFill(
    start_color = "EE1111",
    end_color = "EE1111",
    fill_type = "solid"
)

# Add a two-color scale
ws.conditional_formatting.add(
    "A1:A10",
    ColorScaleRule(
        start_type = "min",
        start_color = "AA0000",
        end_type = "max",
        end_color = "00AA00"
    )
)

# Add a three-color scale
ws.conditional_formatting.add(
    "B1:B10",
    ColorScaleRule(
        start_type = "percentile",
        start_value = 10,
        start_color = "FF0000",
        mid_type = "percentile",
        mid_value = 50,
        mid_color = "0000FF",
        end_type = "percentile",
        end_value = 90,
        end_color = "00AA00"
    )
)

# Formatting baed on a cell comparison

ws.conditional_formatting.add(
    "C2:C10",
    CellIsRule(
        operator = "between",
        formula = ['1', '5'],
        stopIfTrue  = True,
        fill = redFill
    )
)

myFont = Font()
myBorder = Border()
ws.conditional_formatting.add(
    "D1:D10",
    FormulaRule(
        formula = ["E1=0"],
        font = myFont,
        border = myBorder,
        fill = redFill
    )
)

wb.save("conditional-formatting.xlsx")