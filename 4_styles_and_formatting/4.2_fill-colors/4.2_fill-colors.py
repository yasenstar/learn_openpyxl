from openpyxl import Workbook
from openpyxl.styles import PatternFill

workbook = Workbook()
sheet = workbook.active

sheet["A1"] = "Yellow Fill"

# Solid Fill Style
my_fill1 = PatternFill(
    start_color = "FFFF00",
    end_color = "FFFF00",
    fill_type = "solid"
)

sheet["A1"].fill = my_fill1

sheet["B2"] = "Light Up Blue"

# Solid Fill Style
my_fill2 = PatternFill(
    start_color = "0000FF",
    end_color = "FFFFFF",
    fill_type = "mediumGray"
)

sheet["B2"].fill = my_fill2

workbook.save("fill_styles.xlsx")