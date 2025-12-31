from openpyxl import Workbook

workbook = Workbook()
sheet = workbook.active

sheet["A1"], sheet["A2"], sheet["A3"] = 10, 20, 30
sheet["A10"] = "=SUM(A1:A2)" # suggest to keep full capital
sheet["A11"] = "=sum(A1:A2)"
sheet["A12"] = "=AVERAGE(A1:A3)"

sheet["B1"] = '=IF(A1>100, "High", "Low")'

# Using variable for ranges
start_row, end_row = 1, 3
sheet["B2"] = f"=SUM(A{start_row}:A{end_row})"

sheet1 = workbook.create_sheet('Sheet1')
sheet2 = workbook.create_sheet('Sheet2')
sheet1["A1"] = 25
sheet2["B1"] = "=Sheet1!A1 * Sheet1!A1"

workbook.save("formula_example.xlsx")