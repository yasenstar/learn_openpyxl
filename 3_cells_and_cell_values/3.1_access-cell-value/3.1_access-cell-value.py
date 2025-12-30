from openpyxl import load_workbook

workbook = load_workbook("my_workbook.xlsx")
sheet = workbook.active

# Accessing using string index
cell_value = sheet["A2"].value
print(f"Value of A1 is: {cell_value}")

# Accessing using cell() method (row, column)
cell_value1 = sheet.cell(row=2, column=4).value
print(f"Value of D2 is: {cell_value1}")

# Checking for a None value (empty cell)
if sheet["F1"].value is None:
    print("Cell F1 is Empty")

if sheet["A200"].value is None:
    print("Cell A200 is Empty")
else:
    print(f"Value of A200 is: {sheet["A200"].value}")