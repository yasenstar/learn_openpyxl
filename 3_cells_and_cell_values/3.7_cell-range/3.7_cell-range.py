from openpyxl import load_workbook

workbook = load_workbook("my_workbook.xlsx")
sheet = workbook.active

# Iterate through row 1 to 10 and column A to D
for row in sheet.iter_rows(
    min_row = 1,
    max_row = 5,
    min_col = 1,
    max_col = 5
):
    for cell in row:
        print(cell.value, end = "\t")
    print()

# Iterate through column 1 to 4
for col in sheet.iter_cols(
    # min_row = 1,
    max_row = 8,
    # min_col = 1,
    max_col = 2
):
    for cell in col:
        print(cell.value, end="\t")
    print()